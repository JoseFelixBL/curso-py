import openpyxl

wb = openpyxl.load_workbook("planilla.xlsx")

# print(wb.sheetnames)
# print(wb["Hoja1"])

# Para que nos devuelva la hoja que se encuentra activa
# que por defecto siempre será la primera
hoja = wb.active
# print(hoja)

wb.create_sheet("Hoja 3")

print(wb.sheetnames)

# Si quiero seleccionar la "Hoja 3"
hoja3 = wb["Hoja 3"]
# Le puedo cambia el título a la hoja
hoja3.title = "Nuevo título"

# # Ahora queremos saber cuántas filas y columnas tiene mi hoja activa
# print(
#     hoja.max_row,
#     hoja.max_column
# )

# Vamos a seleccionar una celda en particular, p.e.: A1
celda = hoja["A1"]
print(celda)
print(celda.value)
celda.value = "Nombre completo"
print(celda.value)

# Forma "más programable" de acceder a una celda
# en esta forma las filas y las columnas NO PARTEN DESDE UN ÍNDICE 0
# son BASE 1
celda2 = hoja.cell(row=2, column=1)
print(
    celda2.value,
    celda2.row,
    celda2.column,
    celda2.coordinate
)

# Ahora podemos ver los valores de todas las celdas de mi hoja
for fila in range(1, hoja.max_row + 1):
    for columna in range(1, hoja.max_column + 1):
        celda = hoja.cell(row=fila, column=columna)
        print(fila, columna, celda.value, celda.coordinate)

# Cómo obtener una fila en particular, o una columna en particular
columna = hoja["A"]
print(columna)

# Y para inspeccionar los valores de esa Tupla
for celda in columna:
    print(celda.row, celda.column, celda.coordinate, celda.value)

# Y para obtener las filas...
fila = hoja["1"]
print(fila)

for celda in fila:
    print(celda.row, celda.column, celda.coordinate, celda.value)

# Agregar más elementos a las filas
hoja.append([1, 2, 3])
print(hoja.rows)

# No entiendo qué valores son estos #############################
for fila in hoja.rows:
    # print(fila.row, fila.column, fila.coordinate, fila.value)
    print(fila)

# Eliminar columnas o filas
# hoja.delete_cols(indice desde donde quiero eliminar , cantidad de columnas/filas a eliminar-por defecto=1)
hoja.delete_rows(1, 1)

# Crear un nevo archivo Excel que tenga todas las modificaciones hechas
wb.save("nuevo_excel.xlsx")
